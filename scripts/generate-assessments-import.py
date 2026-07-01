#!/usr/bin/env python3
"""Generate assessment import SQL from Microsoft Forms Excel export."""

from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SKILLS_SQL = ROOT / "supabase" / "skills.sql"
RPCS_SQL = ROOT / "supabase" / "assessments-rpcs.sql"


def load_skill_codes() -> list[str]:
    text = SKILLS_SQL.read_text()
    return re.findall(r"\('([a-z]+\d+)','(?:safety|basic|intermediate|advanced|survey)'", text)


def section_of(code: str) -> str:
    if code.startswith("sv"):
        return "survey"
    return {"s": "safety", "b": "basic", "i": "intermediate", "a": "advanced"}[code[0]]


def parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip().replace("\xa0", "").strip()
        if not cleaned:
            return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def effective_date(today, completion) -> date | None:
    parsed = parse_date(today)
    if parsed:
        return parsed
    if isinstance(completion, datetime):
        return completion.date()
    return parse_date(completion)


def find_header(ws) -> int:
    for row in range(1, 15):
        value = ws.cell(row, 10).value
        if value == "SAFETY" or (isinstance(value, str) and value.strip() == "SAFETY"):
            return row
    return 6


def norm_name(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def canonical_sheet_name(sheet_name: str) -> str:
    name = norm_name(sheet_name)
    return re.sub(r"\s*-\s*TERMINATED\s*$", "", name, flags=re.IGNORECASE)


def row_respondent_name(ws, row: int) -> str:
    return norm_name(ws.cell(row, 6).value) or norm_name(ws.cell(row, 5).value)


def row_names_match_sheet(tech_name: str, ws, row: int) -> bool:
    name5 = norm_name(ws.cell(row, 5).value)
    name6 = norm_name(ws.cell(row, 6).value)
    if not name5 and not name6:
        return True
    target = tech_name.lower()
    return any(n and n.lower() == target for n in (name5, name6))


def sql_str(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def sql_num(value: float | int | None) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, float):
        return format(value, ".2f").rstrip("0").rstrip(".")
    return str(value)


def avg_section(raw: dict[str, int], section: str) -> float | None:
    values = [score for code, score in raw.items() if section_of(code) == section]
    return round(sum(values) / len(values), 2) if values else None


def parse_workbook(path: Path, region: str) -> list[dict]:
    codes = load_skill_codes()
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict] = []

    skipped_mismatch: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tech_name = canonical_sheet_name(sheet_name)
        header_row = find_header(ws)
        name_row = header_row + 1
        first_data = header_row + 2

        skill_cols: list[int] = []
        col = 10
        while col <= ws.max_column:
            name = ws.cell(name_row, col).value
            if name is None or str(name).strip() == "":
                break
            skill_cols.append(col)
            col += 1
        skill_count = min(len(skill_cols), len(codes))

        sheet_rows: list[dict] = []
        for row in range(first_data, ws.max_row + 1):
            response_id = ws.cell(row, 1).value
            if not response_id or not str(response_id).strip().isdigit():
                continue

            today = ws.cell(row, 8).value
            completion = ws.cell(row, 3).value
            raw: dict[str, int] = {}
            for index in range(skill_count):
                score = ws.cell(row, skill_cols[index]).value
                if isinstance(score, (int, float)) and 1 <= score <= 5:
                    raw[codes[index]] = int(score)
            if not raw:
                continue

            respondent = row_respondent_name(ws, row)
            if not row_names_match_sheet(tech_name, ws, row):
                skipped_mismatch.append(f"{tech_name} (row has {respondent})")
                continue

            comment = None
            for comment_col in range(ws.max_column, skill_cols[-1], -1):
                candidate = ws.cell(row, comment_col).value
                if (
                    isinstance(candidate, str)
                    and len(candidate.strip()) > 3
                    and "comment" not in candidate.lower()
                    and not candidate.strip().startswith("#")
                ):
                    comment = candidate.strip()
                    break

            sheet_rows.append(
                {
                    "tech": tech_name,
                    "date": effective_date(today, completion),
                    "today_blank": parse_date(today) is None,
                    "fallback_date": effective_date(None, completion),
                    "raw": raw,
                    "comment": comment,
                }
            )

        if not sheet_rows:
            continue

        by_date: OrderedDict[date | None, dict] = OrderedDict()
        for item in sheet_rows:
            by_date[item["date"]] = item
        deduped = list(by_date.values())
        latest = deduped[-1]

        for item in deduped:
            if item is latest and item["today_blank"]:
                rows.append({**item, "mode": "latest"})
            elif item["date"] is not None:
                rows.append({**item, "mode": "dated"})

    wb.close()
    if skipped_mismatch:
        print("Skipped mismatched rows:", ", ".join(skipped_mismatch))
    return rows


def row_to_sql_values(row: dict, region: str) -> str:
    raw_json = json.dumps(row["raw"], separators=(",", ":"))
    return (
        f"({sql_str(row['tech'])}, "
        f"{sql_str(row['mode'])}, "
        f"{sql_str(row['date'].isoformat()) if row.get('date') else 'NULL'}, "
        f"{sql_str(row['fallback_date'].isoformat()) if row.get('fallback_date') else 'NULL'}, "
        f"{sql_num(avg_section(row['raw'], 'safety'))}, "
        f"{sql_num(avg_section(row['raw'], 'basic'))}, "
        f"{sql_num(avg_section(row['raw'], 'intermediate'))}, "
        f"{sql_num(avg_section(row['raw'], 'advanced'))}, "
        f"{sql_num(avg_section(row['raw'], 'survey'))}, "
        f"{sql_str(row['comment'])}, "
        f"'{raw_json}'::jsonb)"
    )


def match_technician_expr(regions: list[str]) -> str:
    if len(regions) == 1:
        return f"app_match_technician_id(v.tech_name, {sql_str(regions[0])})"
    parts = [f"app_match_technician_id(v.tech_name, {sql_str(region)})" for region in regions]
    return "coalesce(" + ", ".join(parts) + ")"


def build_import_sql(rows: list[dict], regions: list[str], source_name: str, file_prefix: str) -> str:
    if not rows:
        raise SystemExit("No assessment rows parsed from workbook.")

    values = ",\n  ".join(row_to_sql_values(row, regions[0]) for row in rows)
    match_expr = match_technician_expr(regions)
    region_label = " + ".join(regions)
    region_filters = " OR ".join(f"t.region = {sql_str(region)}" for region in regions)

    return f"""-- Import {region_label} technician assessments from Microsoft Forms export
-- Source: {source_name}
-- PREREQ: run supabase/assessments-rpcs.sql first — or use {file_prefix}-fix-now.sql (all-in-one).
-- Safe to re-run. No temp tables. No ON CONFLICT.

-- === IMPORT DATA ===
SELECT app_upsert_assessment_for_tech(
  {match_expr},
  CASE
    WHEN v.mode = 'latest' THEN coalesce(
      (SELECT max(a.date)
       FROM public.assessments a
       WHERE a.technician_id = {match_expr}),
      v.fallback_date::date
    )
    ELSE v.assessment_date::date
  END,
  v.safety_avg,
  v.basic_avg,
  v.intermediate_avg,
  v.advanced_avg,
  v.survey_avg,
  v.comment,
  v.raw_scores
)
FROM (VALUES
  {values}
) AS v(tech_name, mode, assessment_date, fallback_date, safety_avg, basic_avg, intermediate_avg, advanced_avg, survey_avg, comment, raw_scores)
WHERE {match_expr} IS NOT NULL;

-- Verify
SELECT 'sample_latest' AS check_name, t.name, t.region, max(a.date) AS latest_date, count(*) AS total
FROM public.technicians t
JOIN public.assessments a ON a.technician_id = t.id
WHERE ({region_filters}) AND t.deleted_at IS NULL
GROUP BY t.id, t.name, t.region
ORDER BY t.region, t.name;

SELECT 'june_2026_count' AS check_name, count(*) AS row_count
FROM public.assessments a
JOIN public.technicians t ON a.technician_id = t.id
WHERE ({region_filters}) AND t.deleted_at IS NULL AND a.date >= '2026-06-01';
"""


def build_fix_now_sql(import_sql: str, region_label: str) -> str:
    marker = "-- === IMPORT DATA ==="
    data_sql = import_sql.split(marker, 1)[1].strip()
    rpcs = RPCS_SQL.read_text()
    return (
        "-- RUN THIS ENTIRE FILE in Supabase SQL Editor (one paste). Then hard-refresh app + sign out/in.\n"
        f"-- Combines dashboard RPC fixes + {region_label} assessment import.\n\n"
        f"{rpcs}\n\n"
        f"{data_sql}\n"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--region", help="Technician region code, e.g. SW or NE")
    parser.add_argument(
        "--regions",
        help="Comma-separated region codes for combined imports, e.g. RM,IM",
    )
    parser.add_argument("--prefix", required=True, help="Output file prefix, e.g. sw-assessments")
    args = parser.parse_args()

    if args.regions:
        regions = [part.strip() for part in args.regions.split(",") if part.strip()]
    elif args.region:
        regions = [args.region]
    else:
        raise SystemExit("Provide --region or --regions")

    rows = parse_workbook(args.workbook, regions[0])
    region_label = " + ".join(regions)
    import_sql = build_import_sql(rows, regions, args.workbook.name, args.prefix)
    out_dir = ROOT / "supabase"
    import_path = out_dir / f"{args.prefix}-import.sql"
    fix_now_path = out_dir / f"{args.prefix}-fix-now.sql"
    import_path.write_text(import_sql)
    fix_now_path.write_text(build_fix_now_sql(import_sql, region_label))

    dated = sum(1 for row in rows if row["mode"] == "dated")
    latest = sum(1 for row in rows if row["mode"] == "latest")
    print(f"Wrote {import_path} ({len(rows)} rows: {dated} dated, {latest} latest)")
    print(f"Wrote {fix_now_path}")


if __name__ == "__main__":
    main()
